"""Generate comprehensive test cases Excel for Legal Chatbot QA testing."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = openpyxl.Workbook()

# ── Styles ────────────────────────────────────────────────────────
HEADER_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="6B21A8", end_color="6B21A8", fill_type="solid")  # violet-800
SECTION_FILL = PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid")  # violet-100
PASS_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")  # green-100
FAIL_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")  # red-100
SKIP_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # yellow-100
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

COLUMNS = ["ID", "Nhóm", "Tên test case", "Bước thực hiện", "Kết quả mong đợi", "Kết quả thực tế", "Pass/Fail", "Ghi chú"]
COL_WIDTHS = [8, 18, 30, 45, 45, 30, 10, 25]


def setup_sheet(ws, title):
    ws.title = title
    ws.sheet_properties.tabColor = "6B21A8"
    for i, (col_name, width) in enumerate(zip(COLUMNS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=i, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"


def add_row(ws, row_num, data, is_section=False):
    for i, val in enumerate(data, 1):
        cell = ws.cell(row=row_num, column=i, value=val)
        cell.alignment = WRAP if i >= 4 else CENTER
        cell.border = THIN_BORDER
        if is_section:
            cell.fill = SECTION_FILL
            cell.font = Font(name="Arial", bold=True, size=10)
        else:
            cell.font = Font(name="Arial", size=10)


# ================================================================
# SHEET 1: Hỏi đáp pháp luật đất đai
# ================================================================
ws1 = wb.active
setup_sheet(ws1, "Hỏi đáp Luật Đất đai")

test_cases_1 = [
    # Section header
    ("", "--- NHÓM 1: CÂU HỎI CƠ BẢN VỀ QUYỀN SỬ DỤNG ĐẤT ---", "", "", "", "", "", ""),

    ("TC-DD-01", "Quyền sử dụng đất",
     "Quyền sử dụng đất của cá nhân gồm những gì?",
     "1. Mở chatbot\n2. Nhập: \"Quyền sử dụng đất của cá nhân gồm những gì?\"\n3. Nhấn gửi",
     "Bot trả lời đầy đủ các quyền: chuyển nhượng, cho thuê, thế chấp, thừa kế, tặng cho, góp vốn... Có trích dẫn điều luật cụ thể.",
     "", "", ""),

    ("TC-DD-02", "Quyền sử dụng đất",
     "Thời hạn sử dụng đất ở là bao lâu?",
     "1. Nhập: \"Thời hạn sử dụng đất ở là bao lâu?\"\n2. Nhấn gửi",
     "Bot trả lời: đất ở được sử dụng ổn định lâu dài (không thời hạn). Có trích dẫn Luật Đất đai.",
     "", "", ""),

    ("TC-DD-03", "Quyền sử dụng đất",
     "Thời hạn sử dụng đất đối với đất chung cư được quy định như thế nào?",
     "1. Nhập: \"Thời hạn sử dụng đất đối với đất chung cư được quy định như thế nào?\"\n2. Nhấn gửi",
     "Bot trả lời về thời hạn sử dụng đất chung cư theo quy định pháp luật hiện hành. Nêu rõ quyền sử dụng đất gắn với sở hữu chung cư.",
     "", "", ""),

    ("TC-DD-04", "Quyền sử dụng đất",
     "Đất nông nghiệp có thời hạn sử dụng bao lâu? Hết hạn thì sao?",
     "1. Nhập: \"Đất nông nghiệp có thời hạn sử dụng bao lâu? Hết hạn thì sao?\"\n2. Nhấn gửi",
     "Bot trả lời: thời hạn 50 năm (có thể gia hạn). Nêu điều kiện gia hạn và thủ tục khi hết hạn.",
     "", "", ""),

    # Section header
    ("", "--- NHÓM 2: CHUYỂN NHƯỢNG & CHUYỂN ĐỔI MỤC ĐÍCH ---", "", "", "", "", "", ""),

    ("TC-DD-05", "Chuyển nhượng đất",
     "Điều kiện để được chuyển nhượng quyền sử dụng đất là gì?",
     "1. Nhập: \"Điều kiện để được chuyển nhượng quyền sử dụng đất là gì?\"\n2. Nhấn gửi",
     "Bot liệt kê đủ điều kiện: có GCN, đất không tranh chấp, không bị kê biên, còn thời hạn sử dụng... Trích dẫn điều luật.",
     "", "", ""),

    ("TC-DD-06", "Chuyển nhượng đất",
     "Hợp đồng chuyển nhượng đất có bắt buộc công chứng không?",
     "1. Nhập: \"Hợp đồng chuyển nhượng đất có bắt buộc công chứng không?\"\n2. Nhấn gửi",
     "Bot trả lời: bắt buộc công chứng hoặc chứng thực. Nêu rõ trường hợp ngoại lệ (nếu có) và hậu quả nếu không công chứng.",
     "", "", ""),

    ("TC-DD-07", "Chuyển mục đích",
     "Tôi có 100m² đất nông nghiệp muốn chuyển sang đất ở tại Quận 7, TP.HCM. Cách tính tiền sử dụng đất dựa trên căn cứ nào?",
     "1. Nhập câu hỏi trên\n2. Nhấn gửi",
     "Bot giải thích về bảng giá đất, giá đất cụ thể, hệ số điều chỉnh. Nêu rõ căn cứ pháp lý (Luật Đất đai, Nghị định hướng dẫn).",
     "", "", "Câu hỏi tính toán cụ thể"),

    ("TC-DD-08", "Chuyển mục đích",
     "Thủ tục chuyển mục đích sử dụng đất từ đất trồng cây sang đất ở gồm mấy bước?",
     "1. Nhập: \"Thủ tục chuyển mục đích sử dụng đất từ đất trồng cây sang đất ở gồm mấy bước?\"\n2. Nhấn gửi",
     "Bot liệt kê từng bước: nộp đơn → thẩm định → quyết định cho phép → nộp tiền SDĐ → cập nhật GCN. Nêu cơ quan tiếp nhận.",
     "", "", "Câu hỏi quy trình"),

    # Section header
    ("", "--- NHÓM 3: TÁCH THỬA, HỢP THỬA ---", "", "", "", "", "", ""),

    ("TC-DD-09", "Tách thửa",
     "Trình tự, thủ tục tách thửa đất thổ cư cho con cái hiện nay gồm bao nhiêu bước? Nộp hồ sơ ở đâu?",
     "1. Nhập câu hỏi trên\n2. Nhấn gửi",
     "Bot liệt kê đầy đủ: chuẩn bị hồ sơ → nộp tại Văn phòng đăng ký đất đai → đo đạc → cấp GCN mới. Nêu giấy tờ cần thiết.",
     "", "", "Câu hỏi quy trình"),

    ("TC-DD-10", "Tách thửa",
     "Diện tích tối thiểu để tách thửa đất ở tại TP.HCM là bao nhiêu m²?",
     "1. Nhập: \"Diện tích tối thiểu để tách thửa đất ở tại TP.HCM là bao nhiêu m²?\"\n2. Nhấn gửi",
     "Bot trả lời theo quy định địa phương (TP.HCM). Nêu rõ khác nhau giữa các quận/huyện nếu có.",
     "", "", "Câu hỏi cụ thể địa phương"),

    ("TC-DD-11", "Hợp thửa",
     "Hai thửa đất liền kề khác chủ sở hữu có thể hợp thửa không?",
     "1. Nhập: \"Hai thửa đất liền kề khác chủ sở hữu có thể hợp thửa không?\"\n2. Nhấn gửi",
     "Bot trả lời: cần chuyển nhượng/tặng cho về cùng chủ trước, sau đó mới hợp thửa. Nêu điều kiện hợp thửa.",
     "", "", ""),

    # Section header
    ("", "--- NHÓM 4: THU HỒI ĐẤT, BỒI THƯỜNG ---", "", "", "", "", "", ""),

    ("TC-DD-12", "Thu hồi đất",
     "Chính quyền phải thông báo trước bao nhiêu ngày khi thu hồi đất?",
     "1. Nhập: \"Chính quyền phải thông báo trước bao nhiêu ngày khi thu hồi đất?\"\n2. Nhấn gửi",
     "Bot trả lời: thời hạn thông báo trước 90-180 ngày tùy trường hợp. Trích dẫn điều luật cụ thể.",
     "", "", ""),

    ("TC-DD-13", "Thu hồi đất",
     "Các trường hợp nào Nhà nước được thu hồi đất không bồi thường?",
     "1. Nhập: \"Các trường hợp nào Nhà nước được thu hồi đất không bồi thường?\"\n2. Nhấn gửi",
     "Bot liệt kê: đất sử dụng trái phép, đất được giao không thu tiền, vi phạm pháp luật đất đai... Trích dẫn luật.",
     "", "", ""),

    ("TC-DD-14", "Bồi thường",
     "Cách tính tiền bồi thường khi Nhà nước thu hồi đất nông nghiệp?",
     "1. Nhập câu hỏi trên\n2. Nhấn gửi",
     "Bot giải thích: bồi thường bằng tiền theo giá đất cụ thể, hỗ trợ ổn định đời sống, hỗ trợ đào tạo nghề. Trích dẫn Nghị định.",
     "", "", "Câu hỏi tính toán"),

    ("TC-DD-15", "Bồi thường",
     "Nếu không đồng ý với quyết định bồi thường thu hồi đất thì khiếu nại ở đâu?",
     "1. Nhập câu hỏi trên\n2. Nhấn gửi",
     "Bot hướng dẫn: khiếu nại lần 1 đến UBND cấp huyện → lần 2 đến UBND cấp tỉnh → khởi kiện tại Tòa án. Nêu thời hạn khiếu nại.",
     "", "", ""),

    # Section header
    ("", "--- NHÓM 5: GIẤY CHỨNG NHẬN (SỔ ĐỎ) ---", "", "", "", "", "", ""),

    ("TC-DD-16", "Sổ đỏ",
     "Thủ tục cấp giấy chứng nhận quyền sử dụng đất lần đầu gồm những gì?",
     "1. Nhập câu hỏi trên\n2. Nhấn gửi",
     "Bot liệt kê hồ sơ cần chuẩn bị, cơ quan tiếp nhận, thời gian giải quyết, lệ phí.",
     "", "", ""),

    ("TC-DD-17", "Sổ đỏ",
     "Sổ đỏ đứng tên 2 vợ chồng, một người muốn bán thì có được không?",
     "1. Nhập câu hỏi trên\n2. Nhấn gửi",
     "Bot trả lời: cần có sự đồng ý của cả hai (tài sản chung). Nêu trường hợp ngoại lệ và cách giải quyết.",
     "", "", "Câu tình huống"),

    ("TC-DD-18", "Sổ đỏ",
     "Mất sổ đỏ thì phải làm gì để xin cấp lại?",
     "1. Nhập: \"Mất sổ đỏ thì phải làm gì để xin cấp lại?\"\n2. Nhấn gửi",
     "Bot hướng dẫn: đăng tin mất → nộp đơn xin cấp lại → cơ quan cấp lại trong 30 ngày. Nêu giấy tờ cần.",
     "", "", ""),

    # Section header
    ("", "--- NHÓM 6: THUẾ, LỆ PHÍ LIÊN QUAN ĐẤT ĐAI ---", "", "", "", "", "", ""),

    ("TC-DD-19", "Thuế đất",
     "Khi chuyển nhượng đất phải đóng những loại thuế, phí gì?",
     "1. Nhập câu hỏi trên\n2. Nhấn gửi",
     "Bot liệt kê: thuế TNCN 2%, lệ phí trước bạ 0.5%, phí công chứng, phí đăng ký biến động. Nêu căn cứ pháp lý.",
     "", "", ""),

    ("TC-DD-20", "Thuế đất",
     "Ai chịu trách nhiệm nộp thuế thu nhập cá nhân khi chuyển nhượng đất - bên bán hay bên mua?",
     "1. Nhập câu hỏi trên\n2. Nhấn gửi",
     "Bot trả lời: bên bán chịu thuế TNCN (2% giá chuyển nhượng). Nhưng thực tế các bên có thể thỏa thuận. Trích dẫn luật.",
     "", "", ""),

    ("TC-DD-21", "Thuế đất",
     "Trường hợp nào được miễn thuế khi chuyển nhượng đất?",
     "1. Nhập câu hỏi trên\n2. Nhấn gửi",
     "Bot liệt kê: chuyển nhượng giữa vợ chồng, cha mẹ-con cái, ông bà-cháu... Nêu điều kiện cụ thể.",
     "", "", ""),

    # Section header
    ("", "--- NHÓM 7: TRANH CHẤP ĐẤT ĐAI ---", "", "", "", "", "", ""),

    ("TC-DD-22", "Tranh chấp",
     "Hàng xóm lấn chiếm đất của tôi thì tôi phải làm gì?",
     "1. Nhập: \"Hàng xóm lấn chiếm đất của tôi thì tôi phải làm gì?\"\n2. Nhấn gửi",
     "Bot hướng dẫn: hòa giải tại UBND xã → khiếu nại lên huyện → khởi kiện tại Tòa án. Nêu chứng cứ cần thu thập.",
     "", "", "Câu tình huống thực tế"),

    ("TC-DD-23", "Tranh chấp",
     "Thời hiệu khởi kiện tranh chấp đất đai là bao lâu?",
     "1. Nhập câu hỏi trên\n2. Nhấn gửi",
     "Bot trả lời về thời hiệu khởi kiện (không áp dụng thời hiệu với quyền sử dụng đất / hoặc 30 năm tùy trường hợp). Trích dẫn luật.",
     "", "", ""),

    ("TC-DD-24", "Tranh chấp",
     "Đất thừa kế chưa có sổ đỏ có được chia không? Thủ tục thế nào?",
     "1. Nhập câu hỏi trên\n2. Nhấn gửi",
     "Bot giải thích: vẫn được chia theo pháp luật thừa kế, cần chứng minh nguồn gốc đất. Nêu thủ tục.",
     "", "", "Câu hỏi phức tạp"),

    # Section header
    ("", "--- NHÓM 8: CÂU HỎI NÂNG CAO / EDGE CASE ---", "", "", "", "", "", ""),

    ("TC-DD-25", "Nâng cao",
     "So sánh quyền của người sử dụng đất được giao không thu tiền và được giao có thu tiền?",
     "1. Nhập câu hỏi trên\n2. Nhấn gửi",
     "Bot so sánh hai hình thức: loại nào được chuyển nhượng, thế chấp, cho thuê. Trình bày rõ ràng.",
     "", "", "Câu so sánh"),

    ("TC-DD-26", "Nâng cao",
     "Người nước ngoài có được mua đất ở Việt Nam không?",
     "1. Nhập câu hỏi trên\n2. Nhấn gửi",
     "Bot trả lời: không được mua đất, chỉ được mua nhà ở (căn hộ chung cư, nhà ở riêng lẻ có điều kiện). Trích dẫn Luật Nhà ở + Luật Đất đai.",
     "", "", ""),

    ("TC-DD-27", "Nâng cao",
     "Đất đang thế chấp ngân hàng có chuyển nhượng được không?",
     "1. Nhập câu hỏi trên\n2. Nhấn gửi",
     "Bot trả lời: cần được ngân hàng đồng ý bằng văn bản hoặc giải chấp trước. Nêu rủi ro nếu mua đất thế chấp.",
     "", "", "Câu tình huống"),

    ("TC-DD-28", "Nâng cao",
     "abc xyz 123 ???",
     "1. Nhập: \"abc xyz 123 ???\"\n2. Nhấn gửi",
     "Bot trả lời lịch sự rằng không hiểu câu hỏi, gợi ý các chủ đề có thể hỏi. KHÔNG hallucinate thông tin pháp luật.",
     "", "", "Test câu vô nghĩa"),

    ("TC-DD-29", "Nâng cao",
     "Cho tôi biết tất cả các loại đất theo Luật Đất đai 2024?",
     "1. Nhập câu hỏi trên\n2. Nhấn gửi",
     "Bot liệt kê phân loại đất theo luật mới nhất (nông nghiệp, phi nông nghiệp, chưa sử dụng + các nhóm con). Có dẫn chiếu luật.",
     "", "", "Câu hỏi rộng"),

    ("TC-DD-30", "Nâng cao",
     "Đất nằm trong quy hoạch treo 10 năm thì quyền lợi của người dân thế nào?",
     "1. Nhập câu hỏi trên\n2. Nhấn gửi",
     "Bot giải thích quyền lợi: vẫn được sử dụng, được chuyển nhượng, quy định về điều chỉnh/hủy quy hoạch. Trích dẫn luật.",
     "", "", "Câu hỏi thời sự"),
]

row = 2
for tc in test_cases_1:
    is_section = tc[0] == "" and tc[1].startswith("---")
    if is_section:
        # Merge section header across all columns
        ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLUMNS))
        cell = ws1.cell(row=row, column=1, value=tc[1].replace("---", "").strip())
        cell.fill = SECTION_FILL
        cell.font = Font(name="Arial", bold=True, size=10, color="6B21A8")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN_BORDER
    else:
        add_row(ws1, row, tc)
    row += 1


# ================================================================
# SHEET 2: Tạo hợp đồng (Contract Creation Flow)
# ================================================================
ws2 = wb.create_sheet()
setup_sheet(ws2, "Tạo hợp đồng")

test_cases_2 = [
    # ── NHÓM 1: KHỞI TẠO ──
    ("", "--- NHÓM 1: KHỞI TẠO HỢP ĐỒNG ---", "", "", "", "", "", ""),

    ("TC-HD-01", "Khởi tạo",
     "Tạo hợp đồng — không chỉ rõ loại",
     "1. Nhập: \"Tôi muốn tạo hợp đồng\"\n2. Nhấn gửi",
     "Bot hỏi: \"Bạn muốn tạo loại hợp đồng nào?\" + hiển thị suggestion chips với 4 loại:\n- Hợp đồng cho thuê đất\n- Hợp đồng chuyển nhượng QSDĐ\n- Hợp đồng giao đất có thu tiền SDĐ\n- Hợp đồng cho thuê khu vực biển",
     "", "", ""),

    ("TC-HD-02", "Khởi tạo",
     "Chọn loại từ suggestion chip",
     "1. Sau TC-HD-01, nhấn chip \"Hợp đồng cho thuê đất\"\n2. Quan sát phản hồi",
     "Bot nhận diện đúng loại, bắt đầu hỏi field đầu tiên (\"Họ và tên bên cho thuê\"). Progress bar xuất hiện: 0/34 trường.",
     "", "", ""),

    ("TC-HD-03", "Khởi tạo",
     "Tạo hợp đồng bằng tên viết tắt / không dấu",
     "1. Nhập: \"tao hop dong chuyen nhuong dat\"\n2. Nhấn gửi",
     "Bot nhận diện đúng \"Hợp đồng chuyển nhượng quyền sử dụng đất\" dù không có dấu tiếng Việt.",
     "", "", "Test normalize tiếng Việt"),

    ("TC-HD-04", "Khởi tạo",
     "Tạo bằng cách nói mô tả ngắn",
     "1. Nhập: \"Tôi muốn cho thuê đất\"\n2. Nhấn gửi",
     "Bot nhận diện ra loại \"Hợp đồng cho thuê đất\" từ ngữ cảnh. Bắt đầu hỏi field.",
     "", "", "Test LLM fallback"),

    ("TC-HD-05", "Khởi tạo",
     "Tạo loại hợp đồng không tồn tại",
     "1. Nhập: \"Tạo hợp đồng bảo hiểm xe máy\"\n2. Nhấn gửi",
     "Bot thông báo chưa hỗ trợ loại này. Liệt kê 4 loại có sẵn trong hệ thống. Suggestions hiện 4 loại.",
     "", "", ""),

    ("TC-HD-06", "Khởi tạo",
     "Tạo từ các keyword khác nhau",
     "1. Thử lần lượt:\n  - \"lập hợp đồng cho thuê đất\"\n  - \"viết hợp đồng giao đất\"\n  - \"làm hợp đồng thuê biển\"",
     "Cả 3 cách diễn đạt đều nhận diện đúng loại hợp đồng tương ứng.",
     "", "", "Test intent keywords"),

    # ── NHÓM 2: HĐ CHO THUÊ ĐẤT (34 fields) ──
    ("", "--- NHÓM 2: HỢP ĐỒNG CHO THUÊ ĐẤT (cho_thue_dat — 34 fields) ---", "", "", "", "", "", ""),

    ("TC-HD-07", "Cho thuê đất",
     "Flow đầy đủ — thông tin Bên cho thuê (7 fields)",
     "1. Nhập: \"Tạo hợp đồng cho thuê đất\"\n2. Bot hỏi lần lượt, trả lời:\n  - Họ tên: \"Nguyễn Văn An\"\n  - Ngày sinh: \"15/03/1980\"\n  - Số CCCD: \"079080012345\"\n  - Ngày cấp CCCD: \"20/06/2021\"\n  - Nơi cấp: \"Cục CSQL CCCD\"\n  - Địa chỉ: \"123 Trần Phú, Q.5, TP.HCM\"\n  - SĐT: \"0901234567\"",
     "Bot ghi nhận từng field, hỏi field tiếp theo. Progress bar: 1/34 → 2/34 → ... → 7/34. Mỗi lần trả lời bot nói \"OK, ghi nhận rồi!\" + hỏi tiếp.",
     "", "", "Test 7 fields bên A"),

    ("TC-HD-08", "Cho thuê đất",
     "Flow đầy đủ — thông tin Bên thuê (7 fields)",
     "3. Tiếp TC-HD-07, trả lời 7 fields bên thuê:\n  - Họ tên: \"Trần Thị Bình\"\n  - Ngày sinh: \"22/08/1990\"\n  - Số CCCD: \"048090067890\"\n  - Ngày cấp: \"15/01/2022\"\n  - Nơi cấp: \"Cục CSQL CCCD\"\n  - Địa chỉ: \"456 Nguyễn Huệ, Q.1, TP.HCM\"\n  - SĐT: \"0987654321\"",
     "Progress bar: 8/34 → 14/34. Bot tiếp tục hỏi phần thông tin đất.",
     "", "", "Test 7 fields bên B"),

    ("TC-HD-09", "Cho thuê đất",
     "Flow đầy đủ — thông tin Thửa đất (7 fields)",
     "4. Tiếp, trả lời:\n  - Địa chỉ thửa đất: \"Lô A12, KCN Tân Bình, TP.HCM\"\n  - Diện tích: \"500 m²\"\n  - Số GCN: \"BT 123456\"\n  - Ngày cấp GCN: \"10/05/2018\"\n  - Nơi cấp GCN: \"UBND TP.HCM\"\n  - Mục đích sử dụng: \"Đất thương mại dịch vụ\"\n  - Mô tả: \"Thửa đất mặt tiền đường 20m\" (optional)",
     "Progress bar: 15/34 → 21/34. Bot hỏi tiếp phần tài chính.",
     "", "", "Có 1 field optional"),

    ("TC-HD-10", "Cho thuê đất",
     "Flow đầy đủ — Tài chính & Thời hạn (10 fields)",
     "5. Tiếp:\n  - Giá thuê: \"50.000.000 VNĐ/tháng\"\n  - Phương thức TT: \"Chuyển khoản\"\n  - Lịch TT: \"Thanh toán trước ngày 5 hàng tháng\"\n  - Đặt cọc: \"100.000.000 VNĐ\" (optional)\n  - Ngày bắt đầu: \"01/04/2026\"\n  - Ngày kết thúc: \"01/04/2031\"\n  - Thời hạn: \"5 năm\"\n  - Gia hạn: \"Tự động gia hạn 1 năm\" (optional)",
     "Progress bar tăng dần. Bot chấp nhận số tiền có dấu chấm phân cách.",
     "", "", "Test số tiền lớn"),

    ("TC-HD-11", "Cho thuê đất",
     "Flow đầy đủ — Điều khoản (6 fields) → Hoàn tất",
     "6. Tiếp:\n  - Điều kiện sử dụng đất: \"Đúng mục đích, không gây ô nhiễm\"\n  - Trách nhiệm bảo trì: \"Bên thuê chịu bảo trì thường xuyên\"\n  - Điều kiện chấm dứt: \"Thông báo trước 3 tháng\"\n  - Giải quyết tranh chấp: \"Thương lượng, hòa giải, tòa án TP.HCM\"\n  - Thỏa thuận khác: \"Không\" (optional)\n7. Quan sát phản hồi cuối cùng",
     "Bot thông báo \"Hợp đồng đã sẵn sàng!\". Tự động tạo PDF. Link tải PDF hiện bên dưới. Progress bar: 34/34.",
     "", "", "End-to-end test"),

    ("TC-HD-12", "Cho thuê đất",
     "Tải và kiểm tra PDF",
     "1. Nhấn link tải PDF từ TC-HD-11\n2. Mở file PDF",
     "File .pdf tải về (KHÔNG phải index.html). Nội dung:\n- Header: CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM\n- Tên: HỢP ĐỒNG CHO THUÊ ĐẤT\n- Thông tin Bên A: Nguyễn Văn An, CCCD 079080012345...\n- Thông tin Bên B: Trần Thị Bình...\n- Thông tin đất: Lô A12...\n- Tiếng Việt hiển thị đúng, không lỗi font",
     "", "", "Regression: PDF font + download"),

    # ── NHÓM 3: HĐ CHUYỂN NHƯỢNG QSDĐ (32 fields) ──
    ("", "--- NHÓM 3: HỢP ĐỒNG CHUYỂN NHƯỢNG QSDĐ (chuyen_nhuong_quyen_su_dung_dat — 32 fields) ---", "", "", "", "", "", ""),

    ("TC-HD-13", "Chuyển nhượng đất",
     "Flow đầy đủ — Bên chuyển nhượng (7 fields)",
     "1. Nhập: \"Làm hợp đồng chuyển nhượng quyền sử dụng đất\"\n2. Trả lời:\n  - Họ tên bên CN: \"Lê Văn Cường\"\n  - Ngày sinh: \"10/12/1975\"\n  - CCCD: \"036075054321\"\n  - Ngày cấp: \"05/03/2020\"\n  - Nơi cấp: \"Cục CSQL CCCD\"\n  - Địa chỉ: \"789 Lê Lợi, Đà Nẵng\"\n  - SĐT: \"0912345678\" (optional)",
     "Bot nhận diện đúng loại. Progress bar: \"Hợp đồng chuyển nhượng quyền sử dụng đất\" — 0/32 → 7/32.",
     "", "", "SĐT là optional"),

    ("TC-HD-14", "Chuyển nhượng đất",
     "Flow đầy đủ — Bên nhận chuyển nhượng (7 fields)",
     "3. Tiếp:\n  - Họ tên bên nhận: \"Phạm Thị Dung\"\n  - Ngày sinh: \"25/07/1988\"\n  - CCCD: \"079088098765\"\n  - Ngày cấp: \"12/09/2021\"\n  - Nơi cấp: \"Cục CSQL CCCD\"\n  - Địa chỉ: \"321 Hùng Vương, Đà Nẵng\"\n  - SĐT: \"0976543210\" (optional)",
     "Progress bar: 8/32 → 14/32.",
     "", "", ""),

    ("TC-HD-15", "Chuyển nhượng đất",
     "Flow đầy đủ — Thông tin thửa đất (8 fields)",
     "4. Tiếp:\n  - Địa chỉ thửa đất: \"Thửa 45, tờ BĐ số 12, xã Hòa Phước, Đà Nẵng\"\n  - Diện tích: \"200 m²\"\n  - Số GCN: \"BĐ 567890\"\n  - Ngày cấp GCN: \"15/08/2015\"\n  - Nơi cấp GCN: \"UBND TP Đà Nẵng\"\n  - Mục đích SDĐ: \"Đất ở tại đô thị\"\n  - Thời hạn SDĐ: \"Lâu dài\"\n  - Mô tả: \"Đất mặt tiền đường 10m\" (optional)",
     "Progress bar: 15/32 → 22/32.",
     "", "", ""),

    ("TC-HD-16", "Chuyển nhượng đất",
     "Flow đầy đủ — Tài chính & Ký kết (10 fields) → Hoàn tất",
     "5. Tiếp:\n  - Giá chuyển nhượng: \"3.500.000.000 VNĐ\"\n  - Đặt cọc: \"350.000.000 VNĐ\" (optional)\n  - Phương thức TT: \"Chuyển khoản ngân hàng\"\n  - Lịch TT: \"Đợt 1: 50% khi ký, Đợt 2: 50% khi sang tên\"\n  - Ngày ký HĐ: \"01/03/2026\"\n  - Nơi ký: \"Văn phòng công chứng Đà Nẵng\"\n  - Ngày hoàn tất: \"01/06/2026\"\n  - Trách nhiệm nộp tiền SDĐ: \"Bên nhận chuyển nhượng\"\n  - Trách nhiệm nộp thuế: \"Mỗi bên chịu thuế theo quy định\"\n  - Thủ tục chuyển đổi GCN: \"UBND TP Đà Nẵng\"",
     "Bot thông báo \"Hợp đồng đã sẵn sàng!\". PDF tạo tự động. Progress bar: 32/32.",
     "", "", "Test giá trị tiền tỷ"),

    ("TC-HD-17", "Chuyển nhượng đất",
     "Kiểm tra nội dung PDF chuyển nhượng",
     "1. Tải PDF từ TC-HD-16\n2. Kiểm tra nội dung",
     "PDF có đủ:\n- Thông tin bên CN: Lê Văn Cường\n- Thông tin bên nhận: Phạm Thị Dung\n- Thông tin đất: 200m², Đà Nẵng\n- Giá: 3.500.000.000 VNĐ\n- Thủ tục chuyển đổi GCN: UBND TP Đà Nẵng\n- Phần ký tên BÊN A / BÊN B",
     "", "", ""),

    # ── NHÓM 4: HĐ GIAO ĐẤT CÓ THU TIỀN SDĐ (45 fields) ──
    ("", "--- NHÓM 4: HỢP ĐỒNG GIAO ĐẤT CÓ THU TIỀN SDĐ (giao_dat_co_thu_tien — 45 fields) ---", "", "", "", "", "", ""),

    ("TC-HD-18", "Giao đất",
     "Khởi tạo — nhận diện tên dài",
     "1. Nhập: \"tạo hợp đồng giao đất có thu tiền sử dụng đất\"\n2. Nhấn gửi",
     "Bot nhận diện đúng loại. Progress bar hiện: 0/45 trường. Bắt đầu hỏi field đầu tiên.",
     "", "", "Tên HĐ dài nhất"),

    ("TC-HD-19", "Giao đất",
     "Flow đầy đủ — Bên giao đất (9 fields, có chức vụ + cơ quan)",
     "2. Trả lời:\n  - Họ tên: \"Nguyễn Minh Tuấn\"\n  - Ngày sinh: \"01/01/1970\"\n  - CCCD: \"036070011111\"\n  - Ngày cấp: \"10/10/2020\"\n  - Nơi cấp: \"Cục CSQL CCCD\"\n  - Địa chỉ: \"UBND Quận Liên Chiểu, Đà Nẵng\"\n  - SĐT: \"0935111222\" (optional)\n  - Chức vụ: \"Chủ tịch UBND Quận\" (optional)\n  - Cơ quan: \"UBND Quận Liên Chiểu\" (optional)",
     "Progress bar: 9/45. Bot hỏi tiếp thông tin bên nhận giao đất. Các field optional có thể bỏ trống.",
     "", "", "Có 3 fields optional"),

    ("TC-HD-20", "Giao đất",
     "Flow đầy đủ — Bên nhận + Thửa đất (17 fields)",
     "3-4. Tiếp:\n  Bên nhận (7 fields): Trần Quốc Bảo, 15/05/1985, ...\n  Thửa đất (11 fields):\n  - Địa chỉ: \"Lô C5, KĐT mới Hòa Xuân\"\n  - Diện tích: \"120 m²\"\n  - Số thửa: \"45\"\n  - Tờ BĐ: \"12\"\n  - Mục đích: \"Đất ở tại đô thị\"\n  - Thời hạn: \"Lâu dài\"\n  - Ranh giới: \"Theo bản đồ trích đo\"",
     "Progress bar: 10/45 → 27/45. Bot hỏi tiếp phần tài chính.",
     "", "", "Nhiều fields đất"),

    ("TC-HD-21", "Giao đất",
     "Flow đầy đủ — Tài chính + Thời hạn + Nghĩa vụ (18 fields) → Hoàn tất",
     "5-6. Tiếp:\n  Tài chính:\n  - Tổng tiền SDĐ: \"2.400.000.000 VNĐ\"\n  - Đặt cọc: \"240.000.000 VNĐ\" (opt)\n  - Phương thức TT: \"Chuyển khoản\"\n  - Lịch TT: \"3 đợt theo tiến độ\"\n  Thời hạn + Bàn giao:\n  - Ngày bắt đầu: \"15/03/2026\"\n  - Ngày bàn giao: \"30/03/2026\"\n  - Căn cứ pháp lý: \"QĐ số 123/QĐ-UBND\"\n  Nghĩa vụ + Điều khoản:\n  - Nghĩa vụ SDĐ: \"Đúng mục đích\"\n  - Nghĩa vụ bên giao: \"Bàn giao đúng hạn\"\n  - Nghĩa vụ bên nhận: \"Nộp tiền SDĐ đúng hạn\"\n  - Điều kiện chấm dứt: \"Vi phạm nghĩa vụ\"\n  - Giải quyết tranh chấp: \"Tòa án Đà Nẵng\"",
     "Bot thông báo hoàn tất. PDF tạo tự động. Progress bar: 45/45. Đây là HĐ nhiều fields nhất (45).",
     "", "", "HĐ phức tạp nhất"),

    # ── NHÓM 5: HĐ CHO THUÊ KHU VỰC BIỂN (37 fields) ──
    ("", "--- NHÓM 5: HỢP ĐỒNG CHO THUÊ KHU VỰC BIỂN (cho_thue_khu_vuc_bien — 37 fields) ---", "", "", "", "", "", ""),

    ("TC-HD-22", "Cho thuê biển",
     "Khởi tạo — nhận diện loại đặc thù",
     "1. Nhập: \"Tạo hợp đồng cho thuê khu vực biển\"\n2. Nhấn gửi",
     "Bot nhận diện đúng loại. Progress bar: 0/37. Bắt đầu hỏi.",
     "", "", "Loại HĐ ít phổ biến"),

    ("TC-HD-23", "Cho thuê biển",
     "Flow đầy đủ — Thông tin 2 bên (14 fields)",
     "2-3. Trả lời:\n  Bên cho thuê (7 fields): \"UBND tỉnh Khánh Hòa\", ...\n  Bên thuê (7 fields): \"Cty TNHH Du lịch Biển Xanh\", ...",
     "Progress bar: 0/37 → 14/37. Bot hỏi tiếp thông tin khu vực biển.",
     "", "", ""),

    ("TC-HD-24", "Cho thuê biển",
     "Flow đầy đủ — Thông tin khu vực biển (11 fields đặc thù)",
     "4. Tiếp:\n  - Vị trí: \"Vịnh Nha Trang, phường Vĩnh Nguyên\"\n  - Diện tích: \"10.000 m²\"\n  - Ranh giới: \"Theo bản đồ đính kèm\"\n  - Tọa độ: \"12.25°N, 109.19°E\"\n  - Xã/phường: \"Vĩnh Nguyên\"\n  - Quận/huyện: \"TP Nha Trang\"\n  - Tỉnh/TP: \"Khánh Hòa\"\n  - Mục đích: \"Kinh doanh du lịch biển\"\n  - Tờ trích lục (optional), Tỷ lệ BĐ (opt), Đơn vị lập (opt), Ngày lập (opt), Đơn vị thẩm định (opt)",
     "Progress bar: 15/37 → 25/37. Có 5 fields bản đồ optional.",
     "", "", "Fields tọa độ, bản đồ"),

    ("TC-HD-25", "Cho thuê biển",
     "Flow đầy đủ — Tài chính + Môi trường → Hoàn tất",
     "5-6. Tiếp:\n  - Giá thuê: \"200.000.000 VNĐ/năm\"\n  - Phương thức TT: \"Chuyển khoản\"\n  - Lịch TT: \"Thanh toán hàng năm\"\n  - Đặt cọc (opt): \"200.000.000 VNĐ\"\n  - Ngày BĐ: \"01/06/2026\"\n  - Ngày KT: \"01/06/2046\"\n  - Thời hạn: \"20 năm\"\n  - Quy hoạch: \"Phù hợp QH biển\"\n  - Bảo vệ môi trường: \"Cam kết không xả thải\"\n  - Bảo vệ bờ biển: \"Không xây dựng sát mép nước\"",
     "Bot thông báo hoàn tất. PDF tạo tự động. Link tải hiện. Progress bar: 37/37.",
     "", "", "Fields môi trường đặc thù"),

    # ── NHÓM 6: HỦY HỢP ĐỒNG ──
    ("", "--- NHÓM 6: HỦY HỢP ĐỒNG ---", "", "", "", "", "", ""),

    ("TC-HD-26", "Hủy HĐ",
     "Hủy hợp đồng đang tạo giữa chừng",
     "1. Tạo HĐ cho thuê đất, điền 5 fields\n2. Nhập: \"hủy\"\n3. Nhấn gửi",
     "Bot xác nhận đã hủy. Mode chuyển về normal. Progress bar biến mất. Data đã nhập bị xóa.",
     "", "", ""),

    ("TC-HD-27", "Hủy HĐ",
     "Hủy bằng các cách viết khác nhau",
     "1. Tạo HĐ mới → Nhập: \"hủy hợp đồng\" → Xác nhận hủy\n2. Tạo HĐ mới → Nhập: \"hủy bỏ\" → Xác nhận\n3. Tạo HĐ mới → Nhập: \"cancel\" → Xác nhận\n4. Tạo HĐ mới → Nhập: \"thoát\" → Xác nhận\n5. Tạo HĐ mới → Nhập: \"bỏ qua\" → Xác nhận",
     "Tất cả 5 cách viết đều hủy thành công.",
     "", "", "6 cancel phrases"),

    ("TC-HD-28", "Hủy HĐ",
     "Tên \"Nguyễn Văn Huy\" KHÔNG hủy HĐ",
     "1. Tạo HĐ, đang điền field \"Họ tên\"\n2. Nhập: \"Nguyễn Văn Huy\"",
     "Bot KHÔNG hủy. Ghi nhận \"Nguyễn Văn Huy\" vào field họ tên. Hỏi field tiếp theo bình thường.",
     "", "", "Regression: tên chứa 'huy'"),

    ("TC-HD-29", "Hủy HĐ",
     "\"chuyển khoản\" KHÔNG hủy HĐ",
     "1. Tạo HĐ, đang điền field \"Phương thức thanh toán\"\n2. Nhập: \"chuyển khoản\"",
     "Bot KHÔNG hủy. Ghi nhận \"chuyển khoản\" vào field. Hỏi field tiếp theo.",
     "", "", "Regression: substring 'huy'"),

    ("TC-HD-30", "Hủy HĐ",
     "Hủy sau khi hoàn tất (state=ready)",
     "1. Hoàn tất tạo HĐ (đã xuất PDF)\n2. Nhập: \"hủy\"",
     "Bot xác nhận hủy. HĐ đã hoàn tất vẫn có thể hủy. Mode chuyển về normal.",
     "", "", ""),

    # ── NHÓM 7: XUẤT PDF & XEM TRƯỚC ──
    ("", "--- NHÓM 7: XUẤT PDF & XEM TRƯỚC ---", "", "", "", "", "", ""),

    ("TC-HD-31", "Xuất PDF",
     "Auto-generate PDF khi hoàn tất",
     "1. Hoàn tất tất cả fields cho bất kỳ HĐ nào\n2. Quan sát phản hồi cuối cùng",
     "Bot thông báo \"Hợp đồng đã sẵn sàng!\". PDF link hiện bên dưới tin nhắn tự động. KHÔNG cần gõ \"xuất pdf\".",
     "", "", ""),

    ("TC-HD-32", "Xuất PDF",
     "Xuất PDF bằng lệnh chat",
     "1. Sau khi HĐ ready\n2. Nhập: \"xuất pdf\"",
     "Bot tạo PDF mới và trả link tải. File .pdf tải được, tiếng Việt đúng.",
     "", "", ""),

    ("TC-HD-33", "Xuất PDF",
     "Tải PDF — KHÔNG tải index.html",
     "1. Nhấn link tải PDF\n2. Kiểm tra file tải về",
     "File tải về có đuôi .pdf, mở được bằng PDF reader. KHÔNG phải file index.html hay HTML.",
     "", "", "Regression: download bug"),

    ("TC-HD-34", "Xuất PDF",
     "Nội dung PDF tiếng Việt chính xác",
     "1. Mở PDF đã tải\n2. Kiểm tra tất cả nội dung",
     "Gồm: header CHXHCNVN, tên HĐ, thông tin 2 bên, thông tin đất, điều khoản, phần ký tên. Tiếng Việt có dấu hiển thị đúng (font DejaVu Serif).",
     "", "", "Regression: font crash"),

    ("TC-HD-35", "Xem trước",
     "Xem trước HĐ dạng HTML",
     "1. Sau khi HĐ ready\n2. Nhập: \"xem hợp đồng\"",
     "Modal popup hiện ra với nội dung HĐ dạng HTML. Có nút \"Xuất PDF\" + nút đóng \"X\".",
     "", "", ""),

    # ── NHÓM 8: SỬA THÔNG TIN ──
    ("", "--- NHÓM 8: SỬA THÔNG TIN SAU KHI HOÀN TẤT ---", "", "", "", "", "", ""),

    ("TC-HD-36", "Sửa field",
     "Sửa một field đã nhập",
     "1. Sau khi HĐ ready\n2. Nhập: \"sửa họ tên bên cho thuê = Nguyễn Văn Bình\"",
     "Bot xác nhận đã sửa thành công. Giá trị mới được cập nhật.",
     "", "", ""),

    ("TC-HD-37", "Sửa field",
     "Xuất PDF sau khi sửa",
     "1. Sau TC-HD-36\n2. Nhập: \"xuất pdf\"",
     "PDF mới chứa thông tin đã sửa (\"Nguyễn Văn Bình\" thay vì tên cũ).",
     "", "", ""),

    ("TC-HD-38", "Sửa field",
     "Sửa field không tồn tại",
     "1. Nhập: \"sửa số hộ chiếu = AB123456\"",
     "Bot thông báo không tìm thấy field. Gợi ý tên field hợp lệ (VD: \"Bạn có thể sửa: họ tên, CCCD, địa chỉ...\").",
     "", "", "Error handling"),

    # ── NHÓM 9: EDGE CASES HỢP ĐỒNG ──
    ("", "--- NHÓM 9: EDGE CASES HỢP ĐỒNG ---", "", "", "", "", "", ""),

    ("TC-HD-39", "Edge case",
     "Tạo HĐ mới khi đang có HĐ cũ chưa xong",
     "1. Tạo HĐ cho thuê đất, điền 5 fields\n2. Nhập: \"Tạo hợp đồng chuyển nhượng đất\"",
     "Bot hỏi xác nhận hủy HĐ cũ hay tiếp tục. HOẶC tự hủy cũ và tạo mới.",
     "", "", ""),

    ("TC-HD-40", "Edge case",
     "Hỏi pháp luật khi đang tạo HĐ",
     "1. Đang điền fields cho HĐ\n2. Nhập: \"Luật đất đai quy định gì về cho thuê đất?\"",
     "Bot nhận ra đây là câu hỏi pháp luật, không phải giá trị field. Trả lời câu hỏi rồi tiếp tục hỏi field tiếp.",
     "", "", "Phân biệt Q&A vs field value"),

    ("TC-HD-41", "Edge case",
     "Điền field với giá trị rất dài",
     "1. Bot hỏi \"Mô tả chi tiết thửa đất\"\n2. Nhập 1 đoạn dài 500+ ký tự mô tả chi tiết",
     "Bot chấp nhận giá trị dài. PDF hiển thị đúng, text wrap không bị tràn.",
     "", "", ""),

    ("TC-HD-42", "Edge case",
     "Điền field optional bằng cách bỏ qua",
     "1. Bot hỏi field optional (VD: \"Mô tả chi tiết thửa đất\")\n2. Nhập: \"bỏ qua\" hoặc \"không\"",
     "Bot chấp nhận bỏ qua field optional. Hỏi field tiếp theo. Progress bar vẫn tăng.",
     "", "", ""),

    ("TC-HD-43", "Edge case",
     "Xuất PDF khi chưa có HĐ",
     "1. Ở mode normal (chưa tạo HĐ)\n2. Nhập: \"xuất pdf\"",
     "Bot thông báo lịch sự: \"Chưa có hợp đồng nào. Bạn muốn tạo loại nào?\" + gợi ý.",
     "", "", ""),

    ("TC-HD-44", "Edge case",
     "Tạo 2 HĐ liên tiếp",
     "1. Tạo HĐ cho thuê đất → hoàn tất → tải PDF\n2. Nhập: \"Tạo hợp đồng chuyển nhượng đất\"\n3. Hoàn tất HĐ thứ 2 → tải PDF",
     "Cả 2 HĐ tạo thành công. 2 file PDF khác nhau. Nội dung đúng cho từng loại.",
     "", "", "Test liên tiếp"),
]

row = 2
for tc in test_cases_2:
    is_section = tc[0] == "" and tc[1].startswith("---")
    if is_section:
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLUMNS))
        cell = ws2.cell(row=row, column=1, value=tc[1].replace("---", "").strip())
        cell.fill = SECTION_FILL
        cell.font = Font(name="Arial", bold=True, size=10, color="6B21A8")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN_BORDER
    else:
        add_row(ws2, row, tc)
    row += 1


# ================================================================
# SHEET 3: Chức năng chung & Edge Cases
# ================================================================
ws3 = wb.create_sheet()
setup_sheet(ws3, "Chức năng chung & Edge Cases")

test_cases_3 = [
    ("", "--- NHÓM 1: SESSION & SIDEBAR ---", "", "", "", "", "", ""),

    ("TC-GN-01", "Session",
     "Tạo phiên hội thoại mới",
     "1. Nhấn nút \"Cuộc hội thoại mới\" trên sidebar\n2. Quan sát giao diện",
     "Chat được reset. Không còn tin nhắn cũ. Suggestions mặc định xuất hiện. Placeholder hiện.",
     "", "", ""),

    ("TC-GN-02", "Session",
     "Gửi tin nhắn đầu tiên — tạo session tự động",
     "1. Ở giao diện mới (chưa có session)\n2. Nhập: \"Xin chào\"\n3. Nhấn gửi",
     "Session mới được tạo. Sidebar hiện session mới với tiêu đề. Bot trả lời chào và gợi ý.",
     "", "", ""),

    ("TC-GN-03", "Session",
     "Chuyển giữa các session cũ",
     "1. Nhấn vào một session cũ trên sidebar\n2. Quan sát giao diện",
     "Tin nhắn cũ được load lại đúng thứ tự. PDF links (nếu có) vẫn hoạt động. Progress bar hiện lại nếu đang tạo HĐ.",
     "", "", ""),

    ("TC-GN-04", "Session",
     "Xóa session",
     "1. Nhấn nút xóa (trash icon) trên một session\n2. Xác nhận xóa",
     "Session bị xóa khỏi sidebar. Tin nhắn trong session bị xóa.",
     "", "", ""),

    ("", "--- NHÓM 2: SUGGESTION CHIPS ---", "", "", "", "", "", ""),

    ("TC-GN-05", "Suggestion chips",
     "Suggestions mặc định khi mới mở",
     "1. Mở chatbot (chưa gửi tin nhắn)\n2. Quan sát phần suggestions",
     "Hiện 3 chip mặc định: \"Tôi có thể hỏi gì về pháp luật?\", \"Tạo hợp đồng thuê nhà\", \"Quyền lợi của người lao động?\"",
     "", "", ""),

    ("TC-GN-06", "Suggestion chips",
     "Nhấn vào suggestion chip",
     "1. Nhấn vào chip \"Tạo hợp đồng thuê nhà\"\n2. Quan sát",
     "Tin nhắn được gửi tự động. Bot phản hồi như khi user tự gõ. Suggestions thay đổi phù hợp với context.",
     "", "", ""),

    ("TC-GN-07", "Suggestion chips",
     "Suggestions sau khi hoàn tất hợp đồng",
     "1. Hoàn tất tạo hợp đồng (đã xuất PDF)\n2. Quan sát suggestions",
     "Hiện chip: \"Tạo hợp đồng mới\", \"Hỏi về luật\" (hoặc tương tự).",
     "", "", ""),

    ("", "--- NHÓM 3: STREAMING & HIỂN THỊ ---", "", "", "", "", "", ""),

    ("TC-GN-08", "Streaming",
     "Phản hồi streaming cho câu hỏi pháp luật",
     "1. Nhập câu hỏi pháp luật (VD: \"Quyền thừa kế theo pháp luật?\")\n2. Quan sát phản hồi",
     "Text hiện ra từng chữ (streaming). Loading dots hiện khi đang chờ. Sau khi xong, text đầy đủ.",
     "", "", ""),

    ("TC-GN-09", "Hiển thị",
     "Hiển thị markdown (bold) trong phản hồi",
     "1. Hỏi một câu pháp luật\n2. Quan sát format phản hồi",
     "Text in đậm (**bold**) hiển thị đúng. Nội dung dễ đọc, có cấu trúc rõ ràng.",
     "", "", ""),

    ("TC-GN-10", "Hiển thị",
     "Tin nhắn user vs bot hiển thị đúng",
     "1. Gửi vài tin nhắn\n2. Quan sát layout",
     "Tin user: bên phải, nền gradient tím-hồng. Tin bot: bên trái, nền trắng, icon bot tím. Avatar hiện đúng.",
     "", "", "Test UI layout"),

    ("", "--- NHÓM 4: ERROR HANDLING ---", "", "", "", "", "", ""),

    ("TC-GN-11", "Error handling",
     "Mất kết nối mạng",
     "1. Tắt wifi/mạng\n2. Gửi tin nhắn\n3. Quan sát",
     "Bot hiện thông báo lỗi: \"Xin lỗi, đã xảy ra lỗi kết nối. Vui lòng thử lại.\" Không crash.",
     "", "", ""),

    ("TC-GN-12", "Error handling",
     "Gửi tin nhắn rỗng",
     "1. Không nhập gì\n2. Nhấn nút gửi",
     "Nút gửi bị disabled hoặc không gửi tin nhắn rỗng. Không có lỗi hiện ra.",
     "", "", ""),

    ("TC-GN-13", "Error handling",
     "Gửi tin nhắn rất dài (>5000 ký tự)",
     "1. Copy-paste một đoạn text rất dài\n2. Nhấn gửi",
     "Bot vẫn xử lý được. Phản hồi bình thường. Không crash hoặc timeout.",
     "", "", ""),

    ("TC-GN-14", "Error handling",
     "Gửi ký tự đặc biệt và emoji",
     "1. Nhập: \"Luật đất đai 🏠 <script>alert('xss')</script>\"\n2. Nhấn gửi",
     "Bot trả lời bình thường (bỏ qua script tag). KHÔNG bị XSS. Emoji hiển thị đúng.",
     "", "", "Test bảo mật XSS"),

    ("TC-GN-15", "Error handling",
     "Nhấn gửi liên tục nhiều lần (spam)",
     "1. Gõ tin nhắn\n2. Nhấn nút gửi 5 lần liên tiếp thật nhanh",
     "Chỉ gửi 1 tin nhắn (nút bị disabled khi đang loading). Không duplicate messages.",
     "", "", ""),

    ("", "--- NHÓM 5: CÂU HỎI PHÁP LUẬT KHÁC (KHÔNG PHẢI ĐẤT ĐAI) ---", "", "", "", "", "", ""),

    ("TC-GN-16", "Luật lao động",
     "Hỏi về quyền lợi người lao động",
     "1. Nhập: \"Người lao động bị sa thải trái pháp luật thì được bồi thường gì?\"\n2. Nhấn gửi",
     "Bot trả lời về quyền lợi: lương, trợ cấp thôi việc, bồi thường... Trích dẫn Bộ luật Lao động.",
     "", "", "Test domain khác"),

    ("TC-GN-17", "Luật dân sự",
     "Hỏi về hợp đồng vay tiền",
     "1. Nhập: \"Lãi suất cho vay tối đa giữa cá nhân theo quy định pháp luật là bao nhiêu?\"\n2. Nhấn gửi",
     "Bot trả lời: không quá 20%/năm khoản tiền vay. Trích dẫn Bộ luật Dân sự.",
     "", "", ""),

    ("TC-GN-18", "Luật doanh nghiệp",
     "Hỏi về thành lập công ty",
     "1. Nhập: \"Vốn điều lệ tối thiểu để thành lập công ty TNHH 1 thành viên?\"\n2. Nhấn gửi",
     "Bot trả lời: không quy định vốn tối thiểu (trừ ngành nghề đặc thù). Nêu rõ nguồn luật.",
     "", "", ""),

    ("TC-GN-19", "Không có dữ liệu",
     "Hỏi lĩnh vực chưa có data",
     "1. Nhập: \"Quy định về xuất nhập khẩu hàng hóa qua cửa khẩu?\"\n2. Nhấn gửi",
     "Nếu chưa có data: Bot thông báo lịch sự \"Hiện chưa có dữ liệu về lĩnh vực này\" + gợi ý các lĩnh vực có sẵn. KHÔNG bịa thông tin.",
     "", "", "Test no-data response"),

    ("TC-GN-20", "Hỏi ngoài lề",
     "Hỏi câu không liên quan pháp luật",
     "1. Nhập: \"Thời tiết hôm nay thế nào?\"\n2. Nhấn gửi",
     "Bot lịch sự giải thích chuyên về pháp luật, gợi ý câu hỏi phù hợp.",
     "", "", "Test off-topic"),
]

row = 2
for tc in test_cases_3:
    is_section = tc[0] == "" and tc[1].startswith("---")
    if is_section:
        ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLUMNS))
        cell = ws3.cell(row=row, column=1, value=tc[1].replace("---", "").strip())
        cell.fill = SECTION_FILL
        cell.font = Font(name="Arial", bold=True, size=10, color="6B21A8")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN_BORDER
    else:
        add_row(ws3, row, tc)
    row += 1


# ── Save ──────────────────────────────────────────────────────────
output = "data/test_cases_legal_chatbot.xlsx"
wb.save(output)
print(f"Saved: {output}")
print(f"Sheets: {wb.sheetnames}")
total = sum(1 for ws in wb.worksheets for r in ws.iter_rows(min_row=2) if r[0].value and r[0].value.startswith("TC-"))
print(f"Total test cases: {total}")
